import msoffcrypto
from openpyxl import load_workbook
import pandas as pd
import io

# Caminho do arquivo Excel criptografado
arquivo = r""

try:
    # Código de descriptografia
    with open(arquivo, 'rb') as f:
        file_content = io.BytesIO(f.read())

    decrypted = io.BytesIO()
    file = msoffcrypto.OfficeFile(file_content)
    file.load_key()  # Troque None pela senha se souber
    file.decrypt(decrypted)

    # Carregar o arquivo Excel
    workbook = load_workbook(filename=decrypted, read_only=False, keep_vba=True, data_only=True)
    print(f"Planilhas disponíveis: {workbook.sheetnames}")

except msoffcrypto.exceptions.DecryptionError:
    print("Não foi possível descriptografar o arquivo. Certifique-se de que a senha está correta.")
except Exception as e:
    print(f"Ocorreu um erro: {e}")
